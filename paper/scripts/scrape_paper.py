#!/usr/bin/env python3
"""
Extract MHC allele strings from a published paper's supplementary data.

Supports:
  - Excel files (.xlsx, .xls)
  - CSV/TSV files
  - Word documents (.docx)
  - Plain text (one allele per line or whitespace-separated)

The script scans all columns/cells for strings that look like MHC allele
names and outputs a validation TSV.

Usage:
    # Single file with known species
    python paper/scripts/scrape_paper.py \
        --input supp_table_S1.xlsx \
        --species "Parus major" \
        --source "PMID:12345678" \
        --output paper/validation/pmid_12345678.tsv

    # Auto-detect species from column headers
    python paper/scripts/scrape_paper.py \
        --input supp_table.csv \
        --source "doi:10.1234/example" \
        --output paper/validation/example.tsv

    # Batch: process all files in a directory
    python paper/scripts/scrape_paper.py \
        --input-dir paper/raw/pmid_12345678/ \
        --species "Gallus gallus" \
        --source "PMID:12345678" \
        --output paper/validation/pmid_12345678.tsv
"""

import argparse
import csv
import re
import sys
from pathlib import Path

SUPPORTED_SUFFIXES = (".xlsx", ".xls", ".csv", ".tsv", ".txt", ".docx")
TOKEN_PATTERN = re.compile(r"[A-Za-z0-9][A-Za-z0-9*:+/_.-]*")
ALLELE_FIELDS_PATTERN = re.compile(r"^\d+(?::\d+)*[A-Za-z]?$")
UNPREFIXED_ALLELE_PATTERN = re.compile(r"^[A-Z][A-Za-z0-9]{0,6}\*[\d:]+[A-Za-z]?$")
NONMAMMALIAN_ALLELE_PATTERN = re.compile(
    r"^(?:UA[A-Z]?|UB[A-Z]?|DA[AB]|DB[AB]|DC[AB]|DR[AB]|BLB|BF)\d?\*[\d:]+$"
)
H2_PATTERN = re.compile(r"^(?:H-?2-[A-Z][A-Za-z0-9]*|H-?2[A-Z][a-z0-9]+)$")
RT1_PATTERN = re.compile(r"^RT1-[A-Za-z0-9.]+$")
MHC_PREFIX_PATTERN = re.compile(r"^Mhc[A-Z][A-Za-z]+-[A-Za-z0-9*:]+$")
HLA_SEROTYPE_PATTERN = re.compile(r"^HLA-[A-Z]\d{1,3}$")
GENE_SUFFIX_PATTERN = re.compile(r"^[A-Z][A-Z0-9]{0,7}$")
GENERIC_MHC_LABEL_PATTERN = re.compile(
    r"^MHC(?:[- ]?(?:CLASS[- ]?)?)?(?:I{1,3}|IA|IB|IIA|IIB|IIIA|IIIB)$",
    re.IGNORECASE,
)
ALL_CAPS_GENE_PREFIXES = {"HLA", "DLA", "SLA", "ELA", "FLA", "MIC"}
STRIP_CHARS = "[](){}<>,;.!?\"'#`"

# Strings that match patterns but are NOT MHC names
FALSE_POSITIVE_PATTERNS = [
    re.compile(r"^\d+$"),  # bare numbers
    re.compile(r"^[A-Z]-\d+$"),  # plate coordinates like A-1
    re.compile(r"^rs\d+$"),  # SNP IDs
    re.compile(r"^chr\d", re.IGNORECASE),  # chromosome names
    re.compile(r"^ENSG\d"),  # Ensembl IDs
    re.compile(r"^LOC\d"),  # NCBI LOC IDs
    re.compile(r"^p\.\w+$"),  # protein changes
]


def clean_candidate_token(token):
    """Strip lightweight wrappers while preserving allele punctuation."""
    return token.strip(STRIP_CHARS)


def has_false_positive_shape(token):
    for fp in FALSE_POSITIVE_PATTERNS:
        if fp.match(token):
            return True
    return False


def looks_like_gene_prefix(prefix):
    """Prefer common immunogenetics prefixes and four-letter species codes."""
    if prefix in ALL_CAPS_GENE_PREFIXES:
        return True
    if len(prefix) == 4 and prefix[0].isupper() and any(ch.islower() for ch in prefix[1:]):
        return all(ch.isalnum() for ch in prefix)
    return (
        len(prefix) >= 4
        and prefix[0].isupper()
        and any(ch.islower() for ch in prefix)
        and any(ch.isupper() for ch in prefix[1:])
        and all(ch.isalnum() for ch in prefix)
    )


def looks_like_gene_suffix(suffix):
    if GENERIC_MHC_LABEL_PATTERN.match(suffix):
        return False
    return bool(GENE_SUFFIX_PATTERN.match(suffix))


def looks_like_mhc_token(token):
    """Check if a single token looks like an MHC allele or gene string."""
    token = clean_candidate_token(token)
    if len(token) < 3 or len(token) > 60:
        return False
    if has_false_positive_shape(token):
        return False
    if GENERIC_MHC_LABEL_PATTERN.match(token):
        return False
    if (
        H2_PATTERN.match(token)
        or RT1_PATTERN.match(token)
        or MHC_PREFIX_PATTERN.match(token)
        or HLA_SEROTYPE_PATTERN.match(token)
        or UNPREFIXED_ALLELE_PATTERN.match(token)
        or NONMAMMALIAN_ALLELE_PATTERN.match(token)
    ):
        return True

    if "-" not in token:
        return False

    prefix, suffix = token.split("-", 1)
    if not prefix or not suffix:
        return False

    if "*" in suffix:
        gene, allele_fields = suffix.split("*", 1)
        return (
            looks_like_gene_prefix(prefix)
            and looks_like_gene_suffix(gene)
            and bool(ALLELE_FIELDS_PATTERN.match(allele_fields))
        )

    return looks_like_gene_prefix(prefix) and looks_like_gene_suffix(suffix)


def split_compound_tokens(token):
    """Split lightweight list-style compounds like A*01:01/A*02:01."""
    if "/" not in token or token.startswith("http"):
        return [token]
    return [part for part in token.split("/") if part]


def looks_like_mhc(s):
    """Check if a string itself looks like an MHC allele/gene token."""
    s = clean_candidate_token(s.strip())
    if not s:
        return False
    return looks_like_mhc_token(s)


def extract_from_text(text):
    """Extract MHC-like tokens from free text."""
    candidates = set()
    for match in TOKEN_PATTERN.finditer(text):
        raw_token = match.group()
        for token in split_compound_tokens(raw_token):
            cleaned = clean_candidate_token(token)
            if looks_like_mhc_token(cleaned):
                candidates.add(cleaned)
    return candidates


def collect_input_files(input_path=None, input_dir=None):
    """Collect scrapeable files from CLI inputs."""
    input_files = []
    if input_path:
        input_files.append(Path(input_path))
    if input_dir:
        directory = Path(input_dir)
        input_files.extend(
            sorted(p for p in directory.iterdir() if p.suffix.lower() in SUPPORTED_SUFFIXES)
        )
    return input_files


def read_excel(path):
    """Read all sheets of an Excel file, return list of cell values."""
    try:
        import openpyxl
    except ImportError:
        print("Requires openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    cells = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells.append(str(cell.value))
    wb.close()
    return cells


def read_csv_tsv(path):
    """Read a CSV or TSV file, return list of cell values."""
    cells = []
    with open(path, newline="", errors="replace") as f:
        # Sniff delimiter
        sample = f.read(4096)
        f.seek(0)
        if "\t" in sample:
            delimiter = "\t"
        elif "," in sample:
            delimiter = ","
        else:
            # Treat as one-per-line
            return [line.strip() for line in f if line.strip()]

        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            cells.extend(row)
    return cells


def read_text(path):
    """Read a plain text file, return list of lines/words."""
    with open(path, errors="replace") as f:
        return [line.strip() for line in f if line.strip()]


def read_docx(path):
    """Read a .docx file and return paragraphs + table cells as text."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
    except ImportError:
        return []

    texts = []
    try:
        with zipfile.ZipFile(path) as z:
            # Read main document
            if "word/document.xml" in z.namelist():
                tree = ET.parse(z.open("word/document.xml"))
                for para in tree.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
                    text = "".join(
                        node.text or ""
                        for node in para.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                    )
                    if text.strip():
                        texts.append(text.strip())
    except Exception:
        pass
    return texts


def process_file(path):
    """Read a file and extract MHC-like strings."""
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        cells = read_excel(path)
    elif suffix in (".csv", ".tsv", ".txt"):
        cells = read_csv_tsv(path)
    elif suffix == ".docx":
        cells = read_docx(path)
    else:
        cells = read_text(path)

    mhc_strings = set()
    for cell in cells:
        mhc_strings.update(extract_from_text(cell))

    return sorted(mhc_strings)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", help="Single input file")
    parser.add_argument("--input-dir", help="Directory of input files")
    parser.add_argument("--species", default="", help="Expected species (latin name)")
    parser.add_argument("--source", required=True, help="Source identifier (PMID or DOI)")
    parser.add_argument(
        "--output", default="-", help="Output TSV file (default: stdout)"
    )
    args = parser.parse_args()

    if not args.input and not args.input_dir:
        parser.error("Provide --input or --input-dir")

    # Collect input files
    input_files = collect_input_files(args.input, args.input_dir)

    if not input_files:
        print("No input files found.", file=sys.stderr)
        return 1

    # Extract from all files
    all_strings = set()
    for path in input_files:
        print(f"Processing {path}...", file=sys.stderr)
        strings = process_file(path)
        print(f"  Found {len(strings)} MHC-like strings", file=sys.stderr)
        all_strings.update(strings)

    print(f"\nTotal unique MHC strings: {len(all_strings)}", file=sys.stderr)

    # Write output
    out = sys.stdout if args.output == "-" else open(args.output, "w")
    writer = csv.DictWriter(
        out,
        fieldnames=["raw_string", "expected_species", "source"],
        delimiter="\t",
    )
    writer.writeheader()
    for s in sorted(all_strings):
        writer.writerow({
            "raw_string": s,
            "expected_species": args.species,
            "source": args.source,
        })

    if args.output != "-":
        out.close()
        print(f"Written to {args.output}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
